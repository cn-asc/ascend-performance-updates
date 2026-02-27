#!/usr/bin/env python3
"""
Metrics-only eval harness: Net MOIC, IRR, and DPI extraction.

Models return JSON as described in EXTRACTION_PROMPT (fund_name, asset_class,
net_irr, net_moic, net_dpi, investment_performance, key_takeaways, business_updates).
Scoring is liberal: we require that ground-truth numerical values appear somewhere
in the JSON (top-level fields or inside the text arrays). Metric naming does not need
to match exactly; extra data that captures ground truth is preferred over narrow
extraction that misses it.
Deterministic eval (no LLM judge). Reports binary match (1/0) per metric, token usage, cost.

Pricing sources:
- OpenAI: https://developers.openai.com/api/docs/pricing (Standard tier)
- Anthropic: https://platform.claude.com/docs/en/about-claude/pricing
"""

import csv
import json
import logging
import os
import random
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

logger = logging.getLogger(__name__)

# Instrumentation: log first N failures (gt non-null but pred null or mismatch)
_FAILURE_LOG_COUNT = 0
_FAILURE_LOG_MAX = 5

from dotenv import load_dotenv

if os.path.exists(".env"):
    load_dotenv()

# -----------------------------------------------------------------------------
# Pricing per 1M tokens (USD). Standard tier for OpenAI; base rates for Anthropic.
# OpenAI: https://developers.openai.com/api/docs/pricing
# Anthropic: https://platform.claude.com/docs/en/about-claude/pricing
# -----------------------------------------------------------------------------
OPENAI_STANDARD = {
    "gpt-5.2": {"input": 1.75, "output": 14.00},
    "gpt-5-mini": {"input": 0.25, "output": 2.00},
    "gpt-5-nano": {"input": 0.05, "output": 0.40},
}
ANTHROPIC_BASE = {
    "claude-sonnet-4-20250514": {"input": 3.00, "output": 15.00}, # Sonnet 4.5
    "claude-haiku-4-5-20251001": {"input": 1.00, "output": 5.00},  # Haiku 4.5 (replaces deprecated 3.5)
    "claude-3-haiku-20240307": {"input": 0.25, "output": 1.25},   # Haiku 3
}
# Google Gemini: https://ai.google.dev/gemini-api/docs/pricing (Standard paid, per 1M tokens)
GEMINI_BASE = {
    "gemini-2.5-flash": {"input": 0.30, "output": 2.50},
}

# Model config: (api_id, provider, display_name).
# OpenAI: gpt-5-mini and gpt-5-nano only support default temperature (1), not 0.
MODELS = [
    ("gpt-5.2", "openai", "GPT-5.2"),
    ("gpt-5-mini", "openai", "GPT-5-mini"),
    ("gpt-5-nano", "openai", "GPT-5-nano"),
    ("claude-sonnet-4-20250514", "anthropic", "Claude Sonnet 4.5"),
    ("claude-haiku-4-5-20251001", "anthropic", "Claude Haiku 4.5"),
    ("claude-3-haiku-20240307", "anthropic", "Claude Haiku 3"),
    ("gemini-2.5-flash", "gemini", "Gemini 2.5 Flash"),
]


def get_pricing(model_id: str, provider: str) -> Tuple[float, float]:
    """Return (input $/1M tokens, output $/1M tokens)."""
    if provider == "openai":
        d = OPENAI_STANDARD.get(model_id) or OPENAI_STANDARD.get("gpt-5.2")
    elif provider == "gemini":
        d = GEMINI_BASE.get(model_id) or GEMINI_BASE.get("gemini-2.5-flash")
    else:
        d = ANTHROPIC_BASE.get(model_id) or ANTHROPIC_BASE.get("claude-3-haiku-20240307")
    return d["input"], d["output"]


# Extraction prompt: three document states — (1) IRR/MOIC/DPI (or TVPI), (2) other performance metrics, (3) no metrics.
EXTRACTION_PROMPT = """You are a highly experienced investment specialist. You are reviewing investment updates that come in a wide array of presentations and formats, in an effort to distill key performance metrics, both qualitative and quantitative. Your task is to extract critical information from investment update documents and return it as structured JSON.

CRITICAL: Always extract fund-level Net IRR, Net MOIC (or TVPI), and Net DPI as separate numeric fields. These are essential for benchmarking.

DOCUMENT TYPES WITH NO FUND-LEVEL METRICS (set net_irr, net_moic, net_dpi to null):
- Investor presentations, pitch decks, investment memos, or announcements (e.g. trip, final close, calendar) that do NOT report fund-level returns.
- If the document clearly has no fund-level performance figures (e.g. "No performance metrics to show", or presentation-only with no fund IRR/MOIC/DPI), set net_irr, net_moic, and net_dpi to null. Do NOT use portfolio-company returns, deal-level IRR/MOIC, or company financials (revenue, EBITDA multiples) as fund-level metrics.

IMPORTANT - When multiple values exist (e.g., fund-level vs. investor-level):
- ALWAYS prioritize values explicitly labeled as "fund level", "fund-level", or "fund level performance"
- NEVER use values labeled for specific investors (e.g., "for a $500K advisory investor", "for advisory investors", "investor-level")
- If you see both fund-level and investor-specific values, extract ONLY the fund-level value
- If only investor-specific values are present, extract those but note the limitation
- ONLY populate net_irr, net_moic, net_dpi when values are explicitly fund-level (or equivalent). Do NOT use deal-level or portfolio-company IRR/MOIC/DPI as fund-level.

For each investment update document provided, extract the following information and return it as a JSON object with these exact keys:

{
  "fund_name": "The exact name of the fund or company",
  "asset_class": "The asset class (e.g., 'Private Equity', 'Venture Capital', 'Real Estate', 'Credit', 'Private Debt', 'Hedge Fund', 'Infrastructure', 'Natural Resources', etc.)",
  "deal_type": "A classification of the investment following menu: [Fund, Direct Investment]"
  "vintage": "The vintage year as a 4-digit number (e.g., 2020, 2021) or null if not found",
  "net_irr": "Net IRR as a number (e.g., 15.5 for 15.5%) or null if not found",
  "net_moic": "Net MOIC or TVPI as a number (e.g., 2.5 for 2.5x) or null if not found",
  "net_dpi": "Net DPI as a number (e.g., 1.2 for 1.2x) or null if not found",
  "performance_summary": "A 1-2 word performance summary from the following menu: [As Expected, Outperforming, Underperforming] - this will be determined after benchmark comparison",
  "investment_performance": [
    "Key metric, return, or performance data point 1",
    "Key metric, return, or performance data point 2",
    "includes quantitative performance metrics, revenue growth, returns, margin expansion,financial data, and new developments",
    "... (as many items as necessary)"
  ],
  "key_takeaways": [
    "Financial performance metric 1",
    "Benchmark comparison 1",
    "Quantitative measure 1",
    "Qualitative measure 1",
    "... (as many items as necessary)"
  ],
  "business_updates": [
    "Strategic update or business development 1",
    "Market condition or commentary 1",
    "... (as many items as necessary)"
  ]
}

IMPORTANT:
- You must not hallucinate under any circumstances. If you are not sure about a piece of information, return null. If you hallucinate, or get any of the information incorrect, someone will die and you will be held responsible. 
- Always extract Asset Class and Vintage explicitly. If not clearly stated, use your best judgment based on context, or use null/empty string if truly unavailable.
- Extract Net IRR, Net MOIC/TVPI, and Net DPI as numeric values (remove % signs, keep as numbers). If not found, use null.
- CRITICAL: For Net IRR, Net MOIC, and Net DPI - if the document shows both fund-level and investor-specific values, you MUST extract the fund-level value. Look for phrases like "fund level", "fund-level", "at the fund level" and ignore values labeled "for a $X investor", "advisory investor", "investor-level", etc.
- PRIORITIZE performance metrics, returns, financial data, and new developments over generic overview information.
- EXCLUDE generic statements like "this fund focuses on X sector" or "the fund invests in Y" - these are not helpful.
- FOCUS on: performance numbers, returns, exits, new investments, portfolio company updates, market conditions, strategic changes.
- Return ONLY valid JSON, no additional text or markdown formatting. Output a single, well-formed JSON object (no trailing prose, no invalid escape sequences, no embedded snippets).
- Use arrays for the three detail sections (investment_performance, key_takeaways, business_updates) - each item should be a clear, concise statement.
- Be thorough but selective - capture important performance and news, skip generic descriptions."""




def _parse_metrics_from_performance_list(perf: List[str]) -> Dict[str, Optional[float]]:
    """Extract first IRR, MOIC/TVPI, DPI from investment_performance—Net or Gross, MOIC or TVPI."""
    out = {"net_irr": None, "net_moic": None, "net_dpi": None}
    for line in perf:
        line = (line or "").strip()
        if not line:
            continue
        # Net IRR / Gross IRR / IRR: 15% or ~32%
        m = re.search(
            r"(?:Net|Gross)\s+IRR[:\s]+~?([\d.]+)\s*%|IRR[:\s]+~?([\d.]+)\s*%|"
            r"Fund Performance:.*?([\d.]+)\s*%\s*Gross IRR|~?([\d.]+)\s*%\s*Gross IRR",
            line, re.I
        )
        if m and out["net_irr"] is None:
            val = next((x for x in (m.group(1), m.group(2), m.group(3), m.group(4)) if x), None)
            if val:
                out["net_irr"] = float(val.replace(",", ""))
        # Net MOIC/TVPI / Gross TVPI / 2.2× TVPI
        m = re.search(
            r"(?:Net\s+)?(?:MOIC|TVPI)[:\s]+([\d.]+)\s*[x×]|(?:Gross\s+)?TVPI[:\s]+([\d.]+)\s*[x×]|"
            r"([\d.]+)\s*[x×]\s*(?:Gross\s+)?(?:TVPI|MOIC)|Fund Performance:\s*([\d.]+)\s*[x×]",
            line, re.I
        )
        if m and out["net_moic"] is None:
            val = next((x for x in (m.group(1), m.group(2), m.group(3), m.group(4)) if x), None)
            if val:
                out["net_moic"] = float(val.replace(",", ""))
        # Net DPI / DPI
        m = re.search(r"Net\s+DPI[:\s]+([\d.]+)\s*[x%]|DPI[:\s]+([\d.]+)\s*[x%]", line, re.I)
        if m and out["net_dpi"] is None:
            val = m.group(1) or m.group(2)
            if val:
                out["net_dpi"] = float(val.replace(",", ""))
    return out


def load_ground_truth(path: str) -> Dict[str, Optional[float]]:
    """
    Load ground truth. Supports two formats:
    1) Explicit: {"net_irr": 15.5, "net_moic": 2.5, "net_dpi": 1.2} (or null)
    2) Existing: {"investment_performance": ["Net IRR: 15%", "Net MOIC: 1.2x", ...]}
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "net_irr" in data or "net_moic" in data or "net_dpi" in data:
        out = {
            "net_irr": data.get("net_irr") if data.get("net_irr") is not None else None,
            "net_moic": data.get("net_moic") if data.get("net_moic") is not None else None,
            "net_dpi": data.get("net_dpi") if data.get("net_dpi") is not None else None,
        }
        if "other_metric_label" in data or "other_metric_value" in data:
            out["other_metric_label"] = data.get("other_metric_label")
            out["other_metric_value"] = normalize_value(data.get("other_metric_value")) if data.get("other_metric_value") is not None else None
        return out
    perf = data.get("investment_performance") or []
    if isinstance(perf, list):
        return _parse_metrics_from_performance_list(perf)
    return {"net_irr": None, "net_moic": None, "net_dpi": None}


def _cell_looks_like_metric_label(cell: Any) -> bool:
    """True if cell looks like a metric name (MOIC, IRR, TVPI, DPI, etc.)."""
    if cell is None:
        return False
    s = str(cell).strip().upper()
    if not s:
        return False
    return (
        "MOIC" in s or "TVPI" in s or "IRR" in s or "DPI" in s
    )


def _cell_looks_like_value(cell: Any) -> bool:
    """True if cell looks like a numeric value (for ground truth)."""
    if cell is None:
        return False
    if isinstance(cell, (int, float)):
        return True
    s = str(cell).strip()
    if not s or s.lower() in ("n/a", "na", "null", "none", ""):
        return False
    try:
        float(s.replace(",", "").replace("%", "").replace("x", "").strip())
        return True
    except ValueError:
        return False


def _parse_one_sheet_metrics(data_rows: List[Tuple[Any, ...]]) -> Dict[str, Optional[float]]:
    """
    Parse metrics from rows 1+ of a single tab (label/value pairs or header+value row).
    Returns one metrics dict with net_irr, net_moic, net_dpi, other_metric_label, other_metric_value.
    """
    metrics = {"net_irr": None, "net_moic": None, "net_dpi": None, "other_metric_label": None, "other_metric_value": None}
    if not data_rows:
        return metrics
    # Try header row + value row: row 0 has labels (irr, moic, tvpi, dpi), row 1 has values
    header = [str(c).strip().lower() if c is not None else "" for c in data_rows[0]]
    if any(h and ("irr" in h or "moic" in h or "tvpi" in h or "dpi" in h) for h in header) and len(data_rows) >= 2:
        val_row = data_rows[1]
        for i, h in enumerate(header):
            if not h:
                continue
            v = normalize_value(val_row[i] if i < len(val_row) else None)
            if "irr" in h and "moic" not in h and "tvpi" not in h:
                metrics["net_irr"] = v
            elif "moic" in h or "tvpi" in h:
                metrics["net_moic"] = v
            elif "dpi" in h:
                metrics["net_dpi"] = v
        return metrics
    # Alternating label/value pairs in column A (and optionally B for value)
    for i in range(0, len(data_rows) - 1, 2):
        label_row = data_rows[i]
        value_row = data_rows[i + 1]
        if not label_row:
            continue
        label_cell = label_row[0] if len(label_row) > 0 else None
        value_cell = value_row[1] if value_row and len(value_row) > 1 and value_row[1] is not None else (value_row[0] if value_row and len(value_row) > 0 else None)
        val = normalize_value(value_cell)
        label_str = str(label_cell).strip() if label_cell else ""
        label_upper = label_str.upper()
        if label_str and _cell_looks_like_metric_label(label_cell):
            if "IRR" in label_upper and "MOIC" not in label_upper and "TVPI" not in label_upper:
                metrics["net_irr"] = val
            elif "MOIC" in label_upper or "TVPI" in label_upper:
                metrics["net_moic"] = val
            elif "DPI" in label_upper:
                metrics["net_dpi"] = val
            else:
                metrics["other_metric_label"] = label_str
                metrics["other_metric_value"] = val
        elif label_str or val is not None:
            metrics["other_metric_label"] = label_str if label_str else None
            metrics["other_metric_value"] = val
    return metrics


def _parse_alternating_label_value_ground_truth(rows: List[Tuple[Any, ...]]) -> Tuple[Dict[str, Dict[str, Optional[float]]], List[Tuple[str, str, Dict[str, Optional[float]]]], List[Dict[str, Optional[float]]]]:
    """
    Parse Excel where each PDF has 2 rows: row 0 = metric label (e.g. MOIC, Gross IRR), row 1 = value.
    No fixed pattern—each document can show a different metric. One document per pair of rows.
    Label from row 2i col 0; value from row 2i+1 col 0 (or col 1 if two columns).
    If row 0 doesn't look like a metric label, treat it as a header and skip it.
    """
    if not rows:
        return {}, [], []
    start = 0
    first_cell = rows[0][0] if rows[0] else None
    if not _cell_looks_like_metric_label(first_cell):
        start = 1
    rows = rows[start:]
    ordered_metrics = []
    for i in range(0, len(rows) - 1, 2):
        label_row = rows[i]
        value_row = rows[i + 1]
        if not label_row:
            continue
        label_cell = label_row[0] if len(label_row) > 0 else None
        # Value: same column as label, or next column if present
        value_cell = value_row[1] if value_row and len(value_row) > 1 and value_row[1] is not None else (value_row[0] if value_row and len(value_row) > 0 else None)
        val = normalize_value(value_cell)
        metrics = {"net_irr": None, "net_moic": None, "net_dpi": None, "other_metric_label": None, "other_metric_value": None}
        label_str = str(label_cell).strip() if label_cell else ""
        label_upper = label_str.upper()
        if label_str and _cell_looks_like_metric_label(label_cell):
            if "IRR" in label_upper and "MOIC" not in label_upper and "TVPI" not in label_upper:
                metrics["net_irr"] = val
            elif "MOIC" in label_upper or "TVPI" in label_upper:
                metrics["net_moic"] = val
            elif "DPI" in label_upper:
                metrics["net_dpi"] = val
            else:
                metrics["other_metric_label"] = label_str
                metrics["other_metric_value"] = val
        elif label_str or val is not None:
            # Honor any other metric (revenue growth, etc.) — state 2
            metrics["other_metric_label"] = label_str if label_str else None
            metrics["other_metric_value"] = val
        # State 3: label and value both empty → metrics stay all null
        # Always add one document per row pair so PDF count = pair count
        ordered_metrics.append(metrics)
    result = {}
    gtpdf_values = []
    for j, m in enumerate(ordered_metrics):
        gtpdf_values.append(("", "", m))
    return result, gtpdf_values, ordered_metrics


def _parse_transposed_ground_truth(rows: List[Tuple[Any, ...]]) -> Tuple[Dict[str, Dict[str, Optional[float]]], List[Tuple[str, str, Dict[str, Optional[float]]]], List[Dict[str, Optional[float]]]]:
    """
    Parse Excel where labels (MOIC, Gross IRR, DPI) are in the first column and values
    are in subsequent columns (one column per document). Row 0 = optional header (doc names).
    Returns (lookup_dict, gtpdf_values, ordered_metrics).
    """
    if len(rows) < 2:
        return {}, [], []
    header = rows[0]
    n_docs = max(0, len(header) - 1)  # columns 1..N = documents
    if n_docs == 0:
        return {}, [], []
    doc_ids = []
    for j in range(1, len(header)):
        c = header[j]
        doc_id = str(c).strip() if c is not None else ""
        doc_ids.append(doc_id)
    ordered_metrics = [{"net_irr": None, "net_moic": None, "net_dpi": None} for _ in range(n_docs)]
    for row in rows[1:]:
        if not row:
            continue
        label_cell = row[0]
        if not _cell_looks_like_metric_label(label_cell):
            continue
        label = str(label_cell).strip().upper()
        for j in range(n_docs):
            col_idx = j + 1
            if col_idx >= len(row):
                continue
            val = normalize_value(row[col_idx])
            if "IRR" in label and "MOIC" not in label and "TVPI" not in label:
                ordered_metrics[j]["net_irr"] = val
            elif "MOIC" in label or "TVPI" in label:
                ordered_metrics[j]["net_moic"] = val
            elif "DPI" in label:
                ordered_metrics[j]["net_dpi"] = val
    result = {}
    gtpdf_values = []
    for j, metrics in enumerate(ordered_metrics):
        doc_id = doc_ids[j] if j < len(doc_ids) else ""
        if doc_id:
            result[doc_id] = metrics
            stem = Path(doc_id).stem
            if stem != doc_id:
                result[stem] = metrics
            gtpdf_values.append((doc_id, stem, metrics))
    return result, gtpdf_values, ordered_metrics


def _metrics_dict_from_json_metrics(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert a JSON 'metrics' object (e.g. {"GTPDF": "...", "Net IRR": 0.103, "Net MOIC": "1.09x"})
    to our standard shape: net_irr, net_moic, net_dpi, other_metric_label, other_metric_value.
    """
    out = {"net_irr": None, "net_moic": None, "net_dpi": None, "other_metric_label": None, "other_metric_value": None}
    for key, val in (raw or {}).items():
        if key in ("GTPDF", "gtpdf", "pdf", "filename"):
            continue
        key_upper = str(key).strip().upper()
        if val is None:
            continue
        v = val
        if isinstance(v, str):
            v = v.strip().rstrip("xX%").strip()
        num = normalize_value(v)  # handles "1.09x", 0.103, etc.
        if ("IRR" in key_upper or key_upper.strip().endswith(" IR")) and "MOIC" not in key_upper and "TVPI" not in key_upper:
            if num is not None and abs(num) <= 2 and (num != 0):
                num = round(num * 100, 2)  # 0.103 -> 10.3
            out["net_irr"] = num
        elif "MOIC" in key_upper or "TVPI" in key_upper:
            out["net_moic"] = num
        elif "DPI" in key_upper:
            out["net_dpi"] = num
        elif out["other_metric_value"] is None and out["other_metric_label"] is None:
            out["other_metric_label"] = str(key).strip()
            out["other_metric_value"] = num
    return out


def load_ground_truth_from_json(json_path: str) -> Tuple[Dict[str, Dict[str, Optional[float]]], List[Tuple[str, str, Dict[str, Optional[float]]]], List[Dict[str, Optional[float]]]]:
    """
    Load multi-document ground truth from a JSON file.
    Expected format: { "DocName": { "metrics": { "GTPDF": "filename.pdf", "Net IRR": 0.1, "Net MOIC": "1.2x", ... } }, ... }
    Returns (lookup_dict, gtpdf_values, ordered_metrics) same as Excel loader.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        return {}, [], []
    result = {}
    gtpdf_values = []
    ordered_metrics = []
    for entry in data.values():
        if not isinstance(entry, dict):
            continue
        metrics_raw = entry.get("metrics") or entry
        if not isinstance(metrics_raw, dict):
            continue
        gtpdf = metrics_raw.get("GTPDF") or metrics_raw.get("gtpdf") or metrics_raw.get("pdf") or metrics_raw.get("filename")
        if not gtpdf:
            continue
        key = str(gtpdf).strip()
        metrics = _metrics_dict_from_json_metrics(metrics_raw)
        ordered_metrics.append(metrics)
        result[key] = metrics
        stem = Path(key).stem if "." in key else key
        if stem != key:
            result[stem] = metrics
        gtpdf_values.append((key, stem, metrics))
    print("Using ground truth JSON (multi-document: GTPDF + metrics per entry).")
    print(f"  Loaded {len(gtpdf_values)} document(s) from JSON.")
    return result, gtpdf_values, ordered_metrics


def load_ground_truth_from_excel(excel_path: str) -> Tuple[Dict[str, Dict[str, Optional[float]]], List[Tuple[str, str, Dict[str, Optional[float]]]], List[Dict[str, Optional[float]]]]:
    """
    Load ground truth from an Excel file. Returns (lookup_dict, gtpdf_values, ordered_metrics).
    Supports three layouts (auto-detected):
    1) Standard: one row per document; header has GTPDF/pdf + metric columns.
    2) Alternating: each PDF has 2 rows — row 0 = metric label (MOIC, Gross IRR, etc.), row 1 = value. No fixed pattern per doc.
    3) Transposed: first column = metric names, next columns = one per document (values).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel ground truth. pip install openpyxl")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    sheets = list(wb.worksheets)

    # One tab per document: each sheet has A1 = document name (GTPDF), rest = that document's metrics
    if len(sheets) >= 2:
        print("Using Excel one-tab-per-document layout (A1 = document name on each tab).")
        result = {}
        gtpdf_values = []
        ordered_metrics = []
        skipped_no_rows = 0
        used_sheet_title = 0
        for ws in sheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                skipped_no_rows += 1
                continue
            raw_key = rows[0][0] if rows[0] else None
            if raw_key is not None:
                key = str(raw_key).strip()
            else:
                key = ""
            if not key:
                # Fallback: use sheet title so we don't skip tabs with empty A1 (e.g. read_only or blank cell)
                key = (ws.title or "").strip() or f"Sheet_{len(gtpdf_values)}"
                used_sheet_title += 1
            metrics = _parse_one_sheet_metrics(rows[1:])
            ordered_metrics.append(metrics)
            result[key] = metrics
            stem = Path(key).stem if "." in key else key
            if stem != key:
                result[stem] = metrics
            gtpdf_values.append((key, stem, metrics))
        wb.close()
        msg = f"  Loaded {len(gtpdf_values)} tab(s) from Excel (A1 = document name per tab)."
        if used_sheet_title:
            msg += f" Used sheet title for {used_sheet_title} tab(s) where A1 was empty."
        if skipped_no_rows:
            msg += f" Skipped {skipped_no_rows} sheet(s) with no rows."
        print(msg)
        return result, gtpdf_values, ordered_metrics

    # Single sheet: standard table (GTPDF column) or alternating/transposed
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}, [], []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    id_cols = ["pdf", "filename", "pdf_filename", "test_case_id", "name", "gtpdf"]
    id_col = None
    for id_name in id_cols:
        if id_name in header:
            id_col = header.index(id_name)
            break
    if id_col is not None:
        # Standard: one row per document; GTPDF column + metric columns.
        print("Using Excel standard layout (GTPDF/filename column + metric columns).")
        def col_index(name: str) -> Optional[int]:
            return header.index(name) if name in header else None
        irr_col = col_index("net_irr") or next((i for i, h in enumerate(header) if h and "irr" in h), None)
        moic_col = col_index("net_moic") or next((i for i, h in enumerate(header) if h and ("moic" in h or "tvpi" in h)), None)
        dpi_col = col_index("net_dpi") or next((i for i, h in enumerate(header) if h and "dpi" in h), None)
        result = {}
        gtpdf_values = []
        ordered_metrics = []
        for row in rows[1:]:
            irr = normalize_value(row[irr_col] if irr_col is not None and irr_col < len(row) else None)
            moic = normalize_value(row[moic_col] if moic_col is not None and moic_col < len(row) else None)
            dpi = normalize_value(row[dpi_col] if dpi_col is not None and dpi_col < len(row) else None)
            metrics = {"net_irr": irr, "net_moic": moic, "net_dpi": dpi}
            ordered_metrics.append(metrics)
            if not row or id_col >= len(row):
                continue
            raw = row[id_col]
            if raw is None:
                continue
            key = str(raw).strip()
            if not key:
                continue
            result[key] = metrics
            stem = Path(key).stem if "." in key else key
            if stem != key:
                result[stem] = metrics
            gtpdf_values.append((key, stem, metrics))
        print(f"  Loaded {len(gtpdf_values)} row(s) from Excel with non-empty GTPDF.")
        wb.close()
        return result, gtpdf_values, ordered_metrics
    # No GTPDF/filename column: try alternating or transposed layouts
    def _value_in_row(row: Tuple[Any, ...]) -> bool:
        if not row or len(row) == 0:
            return False
        return _cell_looks_like_value(row[0]) or (len(row) > 1 and _cell_looks_like_value(row[1]))
    def _value_cells_in_row(row: Tuple[Any, ...]) -> int:
        return sum(1 for c in (row or []) if _cell_looks_like_value(c))
    row0_label = rows[0][0] if rows[0] else None
    row1_label = rows[1][0] if len(rows) > 1 and rows[1] else None
    n_cols = max(len(r) for r in rows) if rows else 0
    if _cell_looks_like_metric_label(row0_label) and len(rows) > 1 and _value_in_row(rows[1]):
        print("Using Excel alternating layout (for each PDF: one row = metric label, next row = value).")
        wb.close()
        return _parse_alternating_label_value_ground_truth(rows)
    if _cell_looks_like_metric_label(row1_label) and len(rows) > 2 and _value_in_row(rows[2]):
        if _value_cells_in_row(rows[2]) >= 2:
            print("Using Excel label-value layout (metric names in first column, one document per column).")
            wb.close()
            return _parse_transposed_ground_truth(rows)
        print("Using Excel alternating layout (for each PDF: one row = metric label, next row = value).")
        wb.close()
        return _parse_alternating_label_value_ground_truth(rows)
    if _cell_looks_like_metric_label(row1_label) and n_cols > 2:
        print("Using Excel label-value layout (metric names in first column, one document per column).")
        wb.close()
        return _parse_transposed_ground_truth(rows)
    wb.close()
    raise ValueError(
        f"Excel must have a GTPDF (or pdf/filename) column for document names, or use label-value layout. Found header: {header}"
    )


def normalize_value(val: Any) -> Optional[float]:
    """Convert to float or None. Handles strings like '15.5' or 'N/A'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if (val == val) else None  # skip NaN
    if isinstance(val, str):
        s = val.strip().lower()
        if s in ("", "n/a", "na", "null", "none"):
            return None
        try:
            return float(s.replace(",", "").rstrip("xX%").strip())
        except ValueError:
            return None
    return None


# Canonical prediction keys (new schema + fallbacks from old schema)
CANONICAL_KEYS = [
    "irr", "moic", "dpi", "tvpi",
    "current_yield", "income_distribution_rate", "gross_annualized_debt_itd_portfolio_yield",
    "total_distributions", "unrealized_value", "closing_date_dec_3_month", "yield", "ytd_whcm", "gmv",
]
# Map other_metric_label (canonicalized) -> canonical key. GT and model labels normalized via canonicalize_label().
OTHER_LABEL_TO_CANONICAL: Dict[str, str] = {
    "current yield": "current_yield",
    "current_yield": "current_yield",
    "income distribution rate": "income_distribution_rate",
    "income_distribution_rate": "income_distribution_rate",
    "distribution rate": "income_distribution_rate",
    "gross annualized debt itd portfolio yield": "gross_annualized_debt_itd_portfolio_yield",
    "gross_annualized_debt_itd_portfolio_yield": "gross_annualized_debt_itd_portfolio_yield",
    "portfolio yield": "gross_annualized_debt_itd_portfolio_yield",
    "annualized debt portfolio yield": "gross_annualized_debt_itd_portfolio_yield",
    "total distributions": "total_distributions",
    "total_distributions": "total_distributions",
    "unrealized value": "unrealized_value",
    "unrealized_value": "unrealized_value",
    "closing date dec 3 month": "closing_date_dec_3_month",
    "closing_date_dec_3_month": "closing_date_dec_3_month",
    "yield": "yield",
    "ytd whcm": "ytd_whcm",
    "ytd_whcm": "ytd_whcm",
    "ytd whc m": "ytd_whcm",
    "gmv": "gmv",
    "spread": "yield",  # spread in bps often reported as yield-like
}


def canonicalize_label(label: str) -> str:
    """Normalize GT and model labels: lowercase, strip punctuation, collapse whitespace. For lookup only."""
    if not label or not str(label).strip():
        return ""
    s = re.sub(r"[^\w\s]", " ", str(label).lower())
    return " ".join(s.split()).strip()


def _normalize_label_for_lookup(label: str) -> str:
    """Normalize a metric label for lookup in OTHER_LABEL_TO_CANONICAL (canonicalize then underscore)."""
    c = canonicalize_label(label)
    if not c:
        return ""
    return c.replace(" ", "_")


def normalize_prediction(pred_json: Dict[str, Any]) -> Dict[str, Optional[float]]:
    """
    Map raw model JSON (old or new schema) to canonical dict with keys in CANONICAL_KEYS.
    Old keys: net_irr, net_moic, net_dpi, net_tvpi, other_metric_label, other_metric_value.
    New keys: irr, moic, dpi, tvpi, current_yield, income_distribution_rate, etc.
    """
    if not pred_json or not isinstance(pred_json, dict):
        return {k: None for k in CANONICAL_KEYS}
    out: Dict[str, Optional[float]] = {k: None for k in CANONICAL_KEYS}
    # Core metrics: new key else old key
    out["irr"] = normalize_value(pred_json.get("irr") or pred_json.get("net_irr"))
    out["moic"] = normalize_value(pred_json.get("moic") or pred_json.get("net_moic"))
    out["dpi"] = normalize_value(pred_json.get("dpi") or pred_json.get("net_dpi"))
    out["tvpi"] = normalize_value(pred_json.get("tvpi") or pred_json.get("net_tvpi"))
    # New-schema named fields
    for key in ["current_yield", "income_distribution_rate", "gross_annualized_debt_itd_portfolio_yield",
                "total_distributions", "unrealized_value", "closing_date_dec_3_month", "yield", "ytd_whcm", "gmv"]:
        if key in pred_json and pred_json[key] is not None:
            out[key] = normalize_value(pred_json[key])
    # Old other_metric_label / other_metric_value -> map to canonical key
    ol = pred_json.get("other_metric_label")
    ov = pred_json.get("other_metric_value")
    if ol is not None and str(ol).strip() and ov is not None:
        norm_label = _normalize_label_for_lookup(str(ol))
        canonical_key = OTHER_LABEL_TO_CANONICAL.get(norm_label)
        if canonical_key and out.get(canonical_key) is None:
            out[canonical_key] = normalize_value(ov)
        # Also try label as "current yield" -> current_yield via space-sep
        if not canonical_key:
            space_key = norm_label.replace("_", " ")
            canonical_key = OTHER_LABEL_TO_CANONICAL.get(space_key)
            if canonical_key and out.get(canonical_key) is None:
                out[canonical_key] = normalize_value(ov)
    return out


def resolve_other_pred_from_canon(canon: Dict[str, Optional[float]], gt_other_label: Optional[str]) -> Optional[float]:
    """Given GT's other_metric_label, return the predicted value from canonical dict."""
    if not gt_other_label or not str(gt_other_label).strip():
        return None
    norm = _normalize_label_for_lookup(str(gt_other_label))
    key = OTHER_LABEL_TO_CANONICAL.get(norm) or OTHER_LABEL_TO_CANONICAL.get(norm.replace("_", " "))
    if key:
        return canon.get(key)
    return None


def unit_normalize_pred(
    pred_val: Optional[float],
    gt_val: Optional[float],
    metric_kind: str,
    label_hint: Optional[str] = None,
) -> Optional[float]:
    """
    Normalize pred value for comparison.
    - Rates: if pred <= 1 and gt > 1, treat pred as fraction -> *100.
    - If gt < 1 and pred > 1, use pred/100 when it reduces absolute error (e.g. 16.3 vs 0.163).
    - bps: if label_hint contains 'bp' or 'bps', convert pred from bps to percent (pred/100).
    - Multiples (moic, dpi, tvpi): no change.
    """
    if pred_val is None:
        return None
    rate_metrics = {"irr", "current_yield", "income_distribution_rate", "gross_annualized_debt_itd_portfolio_yield",
                    "yield", "ytd_whcm", "total_distributions", "unrealized_value", "closing_date_dec_3_month"}
    # bps: raw label/text includes "bp" or "bps" -> pred is in bps, convert to percent
    if label_hint and ("bp" in label_hint.lower() or "bps" in label_hint.lower()):
        pred_val = pred_val / 100.0
    if metric_kind not in rate_metrics:
        return pred_val
    if gt_val is None:
        return pred_val
    # pred as fraction (e.g. 0.103) when gt is percent (e.g. 10.3)
    if abs(gt_val) > 1 and abs(pred_val) <= 1:
        return pred_val * 100
    # pred as percent (e.g. 16.3) when gt is fraction (e.g. 0.163) -> use pred/100 if it reduces error
    if abs(gt_val) < 1 and abs(pred_val) > 1:
        pred_as_fraction = pred_val / 100.0
        if abs(gt_val - pred_as_fraction) < abs(gt_val - pred_val):
            return pred_as_fraction
    return pred_val


def values_match(gt: Optional[float], pred: Optional[float], tolerance: float = 0.01) -> bool:
    """True if both are None, or both are numbers within tolerance."""
    if gt is None and pred is None:
        return True
    if gt is None or pred is None:
        return False
    return abs(gt - pred) <= tolerance


# Regex to find numeric values in text (percent, multiple, or plain float). Liberal extraction for scoring.
_RE_NUMERIC_IN_TEXT = re.compile(
    r"-?\d+(?:,\d{3})*(?:\.\d+)?\s*[%x×]?|-?\d+\.\d+"
)


def _numbers_from_string(s: str) -> List[float]:
    """Extract numeric values from a string (e.g. 'Net IRR: 15.5%' or 'MOIC 1.2x'). Returns list of floats."""
    out: List[float] = []
    if not s or not isinstance(s, str):
        return out
    for m in _RE_NUMERIC_IN_TEXT.finditer(s):
        raw = m.group(0).strip().rstrip("%x×X").replace(",", "").strip()
        try:
            out.append(float(raw))
        except ValueError:
            continue
    return out


def extract_all_performance_numbers(data: Dict[str, Any]) -> List[float]:
    """
    Extract all performance-relevant numbers from the model JSON (EXTRACTION_PROMPT schema).
    Includes top-level net_irr, net_moic, net_dpi and numbers found in
    investment_performance, key_takeaways, business_updates. Liberal: naming does not
    need to match; what matters is that GT values appear somewhere in the data.
    """
    numbers: List[float] = []
    if not data or not isinstance(data, dict):
        return numbers
    for key in ("net_irr", "net_moic", "net_dpi"):
        v = data.get(key)
        n = normalize_value(v)
        if n is not None:
            numbers.append(n)
    for key in ("investment_performance", "key_takeaways", "business_updates"):
        arr = data.get(key)
        if isinstance(arr, list):
            for item in arr:
                if isinstance(item, str):
                    numbers.extend(_numbers_from_string(item))
                elif isinstance(item, (int, float)) and item == item:
                    numbers.append(float(item))
    # Dedupe while preserving order for debugging; use set for match checks
    return numbers


def gt_value_appears_in_set(
    gt_val: Optional[float],
    numbers: List[float],
    tolerance: float = 0.01,
    allow_scale_flex: bool = False,
) -> bool:
    """
    True if gt_val is None or if some value in numbers matches gt_val within tolerance.
    allow_scale_flex: for rate-like metrics (IRR, yield), also accept n*100 or n/100 as match.
    """
    if gt_val is None:
        return True
    for n in numbers:
        if values_match(gt_val, n, tolerance):
            return True
        if allow_scale_flex:
            if values_match(gt_val, n * 100.0, tolerance) or values_match(gt_val, n / 100.0, tolerance):
                return True
    return False


def _strip_json_prefixed_junk(text: str) -> str:
    """Remove common non-JSON prefixes so the first { is the start of the object."""
    if not text:
        return text
    # Strip leading lines that are only "[json]", "json:", "JSON:", etc.
    lines = text.split("\n")
    while lines:
        line = lines[0].strip().lower()
        if not line or line in ("[json]", "json", "json:", "here is the json:", "here is the data:"):
            lines.pop(0)
            continue
        if line.startswith("json:") or line.startswith("```json"):
            lines[0] = lines[0].strip()[lines[0].strip().lower().index("json") + 4:].strip()
            if not lines[0]:
                lines.pop(0)
            continue
        break
    return "\n".join(lines).strip()


def _extract_first_json_object(text: str) -> Optional[str]:
    """Extract the first complete JSON object (handles nested braces). Returns substring or None."""
    text = (text or "").strip()
    if "```json" in text.lower():
        start = text.lower().find("```json") + 7
        end = text.find("```", start)
        if end == -1:
            end = len(text)
        text = text[start:end].strip()
    elif "```" in text:
        start = text.find("```") + 3
        end = text.find("```", start)
        if end == -1:
            end = len(text)
        text = text[start:end].strip()
    text = _strip_json_prefixed_junk(text)
    start = text.find("{")
    if start == -1:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


def _try_parse_json(snippet: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """Try to parse snippet as JSON dict. Returns (data, None) on success, (None, error) on failure."""
    try:
        data = json.loads(snippet)
        if not isinstance(data, dict):
            return None, f"Parsed value is not a dict: {type(data).__name__}"
        return data, None
    except (json.JSONDecodeError, TypeError) as e:
        return None, str(e)


def _repair_invalid_escapes(snippet: str) -> str:
    """Fix Invalid \\escape: in JSON only \\ \" \\/ \\b \\f \\n \\r \\t \\uXXXX are valid. Replace \\+other with the other character."""
    # Match \ not followed by valid escape; replace with just the next char. Allow \u only when +4 hex digits.
    return re.sub(r'\\(?!["\\/bfnrt]|u[0-9a-fA-F]{4})(.)', r'\1', snippet)


def _extract_partial_json(raw: str) -> Dict[str, Any]:
    """When response is truncated (no complete { }), extract net_irr, net_moic, net_dpi from raw text with regex."""
    out: Dict[str, Any] = {}
    # Prefer the first occurrence of each key (in case of duplicated/malformed content)
    for key in ("net_irr", "net_moic", "net_dpi"):
        # Match "net_irr": 26.0 or "net_irr": null
        m = re.search(rf'"{re.escape(key)}"\s*:\s*(-?\d+\.?\d*|null)', raw)
        if m:
            val = m.group(1)
            if val == "null":
                out[key] = None
            else:
                try:
                    out[key] = float(val)
                except ValueError:
                    pass
    return out


def extract_json_from_response(content: str) -> Tuple[Dict[str, Any], Optional[str]]:
    """Parse JSON from model response. Returns (data, parse_error). parse_error is None on success.
    Tries: extract object, parse; fix trailing commas; fix invalid escapes; if no object, extract partial (truncated)."""
    raw = (content or "").strip()
    snippet = _extract_first_json_object(raw)
    if not snippet:
        # Truncated response: try to extract net_irr, net_moic, net_dpi from raw text
        partial = _extract_partial_json(raw)
        if partial:
            logger.warning("No complete JSON object; extracted partial metrics from truncated response (len=%d)", len(raw))
            return partial, "Truncated JSON; extracted partial metrics"
        err = "No JSON object found in response"
        logger.warning("No JSON object found in response (len=%d). Truncated: %s", len(raw), raw[:300])
        return {}, err
    data, err = _try_parse_json(snippet)
    if data is not None:
        return data, None
    # Fix invalid escape sequences (e.g. unescaped backslash or embedded snippet in string)
    if err and ("escape" in err.lower() or "Invalid" in err):
        repaired = _repair_invalid_escapes(snippet)
        data, _ = _try_parse_json(repaired)
        if data is not None:
            return data, None
    # Common LLM mistake: trailing comma before } or ]
    fixed = re.sub(r",\s*}", "}", snippet)
    fixed = re.sub(r",\s*]", "]", fixed)
    data, _ = _try_parse_json(fixed)
    if data is not None:
        return data, None
    # Try repair escapes on the trailing-comma-fixed version
    if err and ("escape" in err.lower() or "Invalid" in err):
        fixed_repaired = _repair_invalid_escapes(fixed)
        data, _ = _try_parse_json(fixed_repaired)
        if data is not None:
            return data, None
    logger.warning(
        "JSON parse failed: %s. Raw truncated: %s. Snippet: %s",
        err, raw[:300], snippet[:500],
    )
    return {}, err


def extract_numbers_from_raw_response(raw_response: str) -> List[float]:
    """When JSON is missing or empty, extract all numeric values from the raw LLM response so we can still score (liberal matching)."""
    return _numbers_from_string(raw_response or "")


def run_openai(
    model_id: str,
    pdf_text: str,
    api_key: Optional[str] = None,
) -> Tuple[Dict[str, Optional[float]], int, int]:
    """Call OpenAI chat completion. Returns (parsed_metrics, input_tokens, output_tokens)."""
    import openai
    client = openai.OpenAI(api_key=api_key or os.getenv("OPENAI_API_KEY"))
    if not client.api_key:
        raise ValueError("OPENAI_API_KEY not set")
    user_content = f"Document text:\n\n{pdf_text[:200000]}"
    # gpt-5-mini and gpt-5-nano only support default temperature (1), not 0
    kwargs = {
        "model": model_id,
        "messages": [
            {"role": "system", "content": EXTRACTION_PROMPT},
            {"role": "user", "content": user_content},
        ],
    }
    if model_id not in ("gpt-5-mini", "gpt-5-nano"):
        kwargs["temperature"] = 0
    response = client.chat.completions.create(**kwargs)
    content = response.choices[0].message.content or "{}"
    usage = getattr(response, "usage", None)
    input_tokens = getattr(usage, "prompt_tokens", 0) or 0
    output_tokens = getattr(usage, "completion_tokens", 0) or 0
    data, parse_error = extract_json_from_response(content)
    parse_ok = parse_error is None
    canon = normalize_prediction(data)
    return canon, input_tokens, output_tokens, content, parse_ok, parse_error, data


def run_anthropic(
    model_id: str,
    pdf_text: str,
    api_key: Optional[str] = None,
) -> Tuple[Dict[str, Optional[float]], int, int]:
    """Call Anthropic messages API. Returns (parsed_metrics, input_tokens, output_tokens)."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key or os.getenv("ANTHROPIC_API_KEY"))
    if not client.api_key:
        raise ValueError("ANTHROPIC_API_KEY not set")
    user_content = f"Document text:\n\n{pdf_text[:200000]}"
    message = client.messages.create(
        model=model_id,
        max_tokens=1024,
        system=EXTRACTION_PROMPT,
        messages=[{"role": "user", "content": user_content}],
    )
    text = ""
    if message.content and isinstance(message.content, list):
        for block in message.content:
            if hasattr(block, "text"):
                text += block.text
    input_tokens = getattr(message, "usage", None)
    if input_tokens:
        in_t = getattr(input_tokens, "input_tokens", 0) or 0
        out_t = getattr(input_tokens, "output_tokens", 0) or 0
    else:
        in_t, out_t = 0, 0
    data, parse_error = extract_json_from_response(text)
    parse_ok = parse_error is None
    canon = normalize_prediction(data)
    return canon, in_t, out_t, text, parse_ok, parse_error, data


def run_gemini(
    model_id: str,
    pdf_text: str,
    api_key: Optional[str] = None,
) -> Tuple[Dict[str, Optional[float]], int, int]:
    """Call Google Gemini API. Uses google-genai SDK if available, else google-generativeai."""
    key = api_key or os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not key:
        raise ValueError("GOOGLE_API_KEY or GEMINI_API_KEY not set")
    user_content = f"Document text:\n\n{pdf_text[:200000]}"
    full_prompt = f"{EXTRACTION_PROMPT}\n\n{user_content}"

    # Prefer new SDK (google-genai); fall back to deprecated google-generativeai if import fails
    try:
        from google.genai import Client
        from google.genai.types import GenerateContentConfig
        client = Client(api_key=key)
        try:
            response = client.models.generate_content(
                model=model_id,
                contents=full_prompt,
                config=GenerateContentConfig(temperature=0),
            )
        except Exception as e:
            _raise_gemini_quota_if_429(e)
        text = (getattr(response, "text", None) or "").strip()
        usage = getattr(response, "usage_metadata", None)
        in_t = int(getattr(usage, "prompt_token_count", 0) or 0)
        out_t = int(getattr(usage, "candidates_token_count", 0) or getattr(usage, "output_token_count", 0) or 0)
    except ImportError:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", FutureWarning)
            import google.generativeai as genai
        genai.configure(api_key=key)
        model = genai.GenerativeModel(model_id)
        try:
            response = model.generate_content(
                full_prompt,
                generation_config=genai.GenerationConfig(temperature=0),
            )
        except Exception as e:
            _raise_gemini_quota_if_429(e)
        text = (getattr(response, "text", None) or "").strip()
        usage = getattr(response, "usage_metadata", None)
        in_t = int(getattr(usage, "prompt_token_count", 0) or 0)
        out_t = int(getattr(usage, "candidates_token_count", 0) or getattr(usage, "output_token_count", 0) or 0)

    data, parse_error = extract_json_from_response(text)
    parse_ok = parse_error is None
    canon = normalize_prediction(data)
    return canon, in_t, out_t, text, parse_ok, parse_error, data


def _raise_gemini_quota_if_429(e: Exception) -> None:
    err_str = str(e)
    if "429" in err_str or "quota" in err_str.lower() or "exceeded" in err_str.lower():
        raise RuntimeError(
            "Gemini quota exceeded (free tier limit or rate limit). "
            "Enable billing at https://ai.google.dev or retry later. See https://ai.google.dev/gemini-api/docs/rate-limits"
        ) from e
    raise


def run_model(
    model_id: str,
    provider: str,
    pdf_text: str,
) -> Tuple[Dict[str, Optional[float]], int, int, float, str, bool, Optional[str], Dict[str, Any]]:
    """Run one model. Returns (canonical_metrics, input_tokens, output_tokens, cost_usd, raw_response, parse_ok, parse_error, raw_parsed_data)."""
    raw = ""
    parse_ok = False
    parse_error: Optional[str] = "unknown"
    data: Dict[str, Any] = {}
    if provider == "openai":
        canon, in_t, out_t, raw, parse_ok, parse_error, data = run_openai(model_id, pdf_text)
    elif provider == "gemini":
        canon, in_t, out_t, raw, parse_ok, parse_error, data = run_gemini(model_id, pdf_text)
    else:
        canon, in_t, out_t, raw, parse_ok, parse_error, data = run_anthropic(model_id, pdf_text)
    in_per_m, out_per_m = get_pricing(model_id, provider)
    cost = (in_t / 1_000_000 * in_per_m) + (out_t / 1_000_000 * out_per_m)
    return canon, in_t, out_t, cost, raw, parse_ok, parse_error, data


# Keyword pattern for --debug-dump-text (lines containing any of these)
_DEBUG_KEYWORD_RE = re.compile(r"irr|moic|dpi|tvpi|yield|bps|ytd", re.I)


_PROJECTED_DOC_KEYWORDS = ("estimate", "projection", "forecast")


def _is_projected_document(pdf_text: str, test_case_id: str) -> bool:
    """True if document is explicitly labeled as estimate/projection/forecast."""
    combined = f"{(pdf_text or '').lower()} {(test_case_id or '').lower()}"
    return any(kw in combined for kw in _PROJECTED_DOC_KEYWORDS)


def compute_score_from_gt_and_matches(
    gt: Dict[str, Any],
    irr_match: int,
    moic_match: int,
    dpi_match: int,
    other_match: int,
) -> Tuple[bool, bool, bool, bool, int, int, Union[str, float], bool, int]:
    """Compute in-scope flags and overall score from GT and match values. In-scope is from GT only."""
    irr_in_scope = gt.get("net_irr") is not None
    moic_in_scope = gt.get("net_moic") is not None
    dpi_in_scope = gt.get("net_dpi") is not None
    other_in_scope = gt.get("other_metric_value") is not None
    score_denom = (1 if irr_in_scope else 0) + (1 if moic_in_scope else 0) + (1 if dpi_in_scope else 0) + (1 if other_in_scope else 0)
    no_metrics_in_scope = score_denom == 0
    if no_metrics_in_scope:
        return irr_in_scope, moic_in_scope, dpi_in_scope, other_in_scope, 0, 0, "", True, 1
    score_num = (irr_match if irr_in_scope else 0) + (moic_match if moic_in_scope else 0) + (dpi_match if dpi_in_scope else 0) + (other_match if other_in_scope else 0)
    overall_score = round(score_num / score_denom, 4)
    return irr_in_scope, moic_in_scope, dpi_in_scope, other_in_scope, score_denom, score_num, overall_score, False, 0


def evaluate_one(
    pdf_path: str,
    ground_truth_path: Optional[str] = None,
    ground_truth_dict: Optional[Dict[str, Optional[float]]] = None,
    test_case_id: Optional[str] = None,
    tolerance: float = 0.01,
    metrics_in_scope: Optional[List[str]] = None,
    debug_dump_text: bool = False,
    allow_projected_metrics: bool = False,
) -> List[Dict[str, Any]]:
    """
    Run all models on one PDF and compare to ground truth.
    Provide either ground_truth_path (JSON) or ground_truth_dict (e.g. from Excel).
    metrics_in_scope: if set, only these canonical metric names count toward score (e.g. irr,moic,dpi,current_yield).
    debug_dump_text: if True, write extracted text to debug_text/<test_case_id>.txt and print keyword hits.
    allow_projected_metrics: if False, documents labeled estimate/projection/forecast are excluded from score_denom.
    Returns list of row dicts for CSV.
    """
    from pdf_processor import extract_text_from_pdf

    if ground_truth_dict is None and ground_truth_path is None:
        raise ValueError("Provide either ground_truth_path or ground_truth_dict")
    if ground_truth_dict is None:
        gt = load_ground_truth(ground_truth_path)
    else:
        gt = ground_truth_dict
    if test_case_id is None:
        test_case_id = Path(pdf_path).stem

    pdf_text = extract_text_from_pdf(pdf_path)
    if not pdf_text:
        raise ValueError(f"No text extracted from PDF: {pdf_path}")
    if debug_dump_text:
        os.makedirs("debug_text", exist_ok=True)
        out_path = os.path.join("debug_text", f"{test_case_id}.txt")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(pdf_text)
        print(f"  [debug] Wrote extracted text to {out_path}")
        keyword_lines = [line.strip() for line in pdf_text.splitlines() if line.strip() and _DEBUG_KEYWORD_RE.search(line)]
        for i, line in enumerate(keyword_lines[:30]):
            print(f"  [debug keyword {i+1}] {line[:120]}{'...' if len(line) > 120 else ''}")
        if len(keyword_lines) > 30:
            print(f"  [debug] ... and {len(keyword_lines) - 30} more keyword lines")
    # Coatue Asia: print lines containing "estimate" or "%" to verify if metric is projected
    if "Coatue" in (test_case_id or ""):
        estimate_pct_lines = [line.strip() for line in pdf_text.splitlines() if line.strip() and ("estimate" in line.lower() or "%" in line)]
        print(f"  [Coatue] Lines containing 'estimate' or '%' ({len(estimate_pct_lines)}):")
        for i, line in enumerate(estimate_pct_lines[:25]):
            print(f"    {i+1}: {line[:130]}{'...' if len(line) > 130 else ''}")
        if len(estimate_pct_lines) > 25:
            print(f"    ... and {len(estimate_pct_lines) - 25} more")
    projected = _is_projected_document(pdf_text, test_case_id or "")
    rows = []
    for model_id, provider, display_name in MODELS:
        parse_ok = True
        parse_error: Optional[str] = None
        try:
            canon, in_t, out_t, cost, raw_response, parse_ok, parse_error, data = run_model(model_id, provider, pdf_text)
        except Exception as e:
            canon = {k: None for k in CANONICAL_KEYS}
            data = {}
            in_t, out_t, cost = 0, 0, 0.0
            raw_response = ""
            parse_ok = False
            parse_error = str(e)
            print(f"  Error {display_name}: {e}")
        # Liberal matching: GT numerical values can appear anywhere in the JSON or in the raw response text. When no JSON is found (parse_error), we still extract numbers from raw text so we can score.
        numbers = extract_all_performance_numbers(data)
        raw_numbers = extract_numbers_from_raw_response(raw_response)
        numbers = numbers + raw_numbers
        irr_match = 1 if gt.get("net_irr") is None else (1 if gt_value_appears_in_set(gt.get("net_irr"), numbers, tolerance, allow_scale_flex=True) else 0)
        moic_match = 1 if gt.get("net_moic") is None else (1 if gt_value_appears_in_set(gt.get("net_moic"), numbers, tolerance, allow_scale_flex=False) else 0)
        dpi_match = 1 if gt.get("net_dpi") is None else (1 if gt_value_appears_in_set(gt.get("net_dpi"), numbers, tolerance, allow_scale_flex=False) else 0)
        other_gt_val = gt.get("other_metric_value")
        other_match = 1 if other_gt_val is None else (1 if gt_value_appears_in_set(other_gt_val, numbers, tolerance, allow_scale_flex=True) else 0)
        # Pred values for CSV display (from canonical dict when available)
        irr_pred_raw = canon.get("irr")
        moic_pred_raw = canon.get("moic")
        dpi_pred_raw = canon.get("dpi")
        irr_pred_norm = unit_normalize_pred(irr_pred_raw, gt.get("net_irr"), "irr")
        moic_pred_norm = unit_normalize_pred(moic_pred_raw, gt.get("net_moic"), "moic")
        dpi_pred_norm = unit_normalize_pred(dpi_pred_raw, gt.get("net_dpi"), "dpi")
        other_pred_val = resolve_other_pred_from_canon(canon, gt.get("other_metric_label"))
        other_metric_kind = "yield"
        _ol = gt.get("other_metric_label")
        _ol_str = str(_ol) if _ol else ""
        if _ol:
            _norm = _normalize_label_for_lookup(_ol_str)
            _key = OTHER_LABEL_TO_CANONICAL.get(_norm) or OTHER_LABEL_TO_CANONICAL.get(canonicalize_label(_ol_str).replace(" ", "_"))
            if _key:
                other_metric_kind = _key
        if other_gt_val is not None and other_pred_val is not None:
            other_pred_val = unit_normalize_pred(other_pred_val, other_gt_val, other_metric_kind, label_hint=_ol_str)
        (irr_in_scope, moic_in_scope, dpi_in_scope, other_in_scope, score_denom, score_num, overall_score, no_metrics_in_scope, overall_score_is_na) = compute_score_from_gt_and_matches(
            gt, irr_match, moic_match, dpi_match, other_match
        )
        metrics_in_denom = []
        if irr_in_scope:
            metrics_in_denom.append("irr")
        if moic_in_scope:
            metrics_in_denom.append("moic")
        if dpi_in_scope:
            metrics_in_denom.append("dpi")
        if other_in_scope:
            metrics_in_denom.append(other_metric_kind)
        # Build grading explanation for tracing: exactly what was graded and why
        def _metric_reason(name: str, gt_v: Any, match: int) -> str:
            if gt_v is None:
                return f"{name}: not in scope"
            if match:
                return f"{name}: GT={gt_v} → found in response (match=1)"
            return f"{name}: GT={gt_v} → not in response (match=0)"
        grading_parts = []
        if irr_in_scope:
            grading_parts.append(_metric_reason("irr", gt.get("net_irr"), irr_match))
        if moic_in_scope:
            grading_parts.append(_metric_reason("moic", gt.get("net_moic"), moic_match))
        if dpi_in_scope:
            grading_parts.append(_metric_reason("dpi", gt.get("net_dpi"), dpi_match))
        if other_in_scope:
            grading_parts.append(_metric_reason(other_metric_kind, other_gt_val, other_match))
        if no_metrics_in_scope:
            grading_parts.append("no_metrics_in_scope; overall_score=N/A")
        else:
            grading_parts.append(f"score_denom={score_denom}, score_num={score_num}, overall_score={overall_score}")
        grading_explanation = "; ".join(grading_parts)
        # Truncate numbers list for display (full list in trace JSON)
        numbers_preview = numbers[:50] if len(numbers) > 50 else numbers
        numbers_preview_str = str(numbers_preview) + (" ..." if len(numbers) > 50 else "")
        # Instrumentation: log first N failures with GT expectations and pred (normalized)
        global _FAILURE_LOG_COUNT
        if _FAILURE_LOG_COUNT < _FAILURE_LOG_MAX and score_denom > 0 and overall_score != "" and overall_score < 1.0:
            _FAILURE_LOG_COUNT += 1
            gt_metrics_present = {}
            if gt.get("net_irr") is not None:
                gt_metrics_present["irr"] = gt.get("net_irr")
            if gt.get("net_moic") is not None:
                gt_metrics_present["moic"] = gt.get("net_moic")
            if gt.get("net_dpi") is not None:
                gt_metrics_present["dpi"] = gt.get("net_dpi")
            if other_gt_val is not None:
                gt_metrics_present["other"] = (gt.get("other_metric_label"), other_gt_val)
            pred_norm = {"irr": irr_pred_norm, "moic": moic_pred_norm, "dpi": dpi_pred_norm, "other": other_pred_val}
            logger.warning(
                "[eval failure %d/%d] test_case_id=%r model_id=%r\n  GT metrics present: %r\n  pred (after unit_normalize): %r\n  metrics_in_denom: %r\n  raw_response_truncated=%r\n  parsed_canon=%r",
                _FAILURE_LOG_COUNT, _FAILURE_LOG_MAX, test_case_id, model_id,
                gt_metrics_present,
                pred_norm,
                metrics_in_denom,
                (raw_response[:400] + "..." if len(raw_response) > 400 else raw_response),
                {k: canon.get(k) for k in CANONICAL_KEYS if canon.get(k) is not None},
            )
        rows.append({
            "test_case_id": test_case_id,
            "model_id": model_id,
            "model_display_name": display_name,
            "provider": provider,
            "input_tokens": in_t,
            "output_tokens": out_t,
            "cost_usd": round(cost, 6),
            "parse_ok": parse_ok,
            "parse_error": parse_error if parse_error is not None else "",
            "overall_score": overall_score,
            "overall_score_is_na": overall_score_is_na,
            "irr_gt": gt.get("net_irr") if gt.get("net_irr") is not None else "",
            "irr_pred": irr_pred_raw if irr_pred_raw is not None else "",
            "irr_match": irr_match,
            "moic_gt": gt.get("net_moic") if gt.get("net_moic") is not None else "",
            "moic_pred": moic_pred_raw if moic_pred_raw is not None else "",
            "moic_match": moic_match,
            "dpi_gt": gt.get("net_dpi") if gt.get("net_dpi") is not None else "",
            "dpi_pred": dpi_pred_raw if dpi_pred_raw is not None else "",
            "dpi_match": dpi_match,
            "other_metric_label_gt": gt.get("other_metric_label") or "",
            "other_metric_value_gt": other_gt_val if other_gt_val is not None else "",
            "other_metric_value_pred": other_pred_val if other_pred_val is not None else "",
            "other_metric_match": other_match,
            "irr_in_scope": irr_in_scope,
            "moic_in_scope": moic_in_scope,
            "dpi_in_scope": dpi_in_scope,
            "other_in_scope": other_in_scope,
            "no_metrics_in_scope": no_metrics_in_scope,
            "grading_explanation": grading_explanation,
            "numbers_extracted_preview": numbers_preview_str,
            "raw_response": raw_response,
            "parsed_data": data,
            "numbers_extracted": numbers,
        })
    return rows


def _sanitize_metric_column_name(label: str) -> str:
    """Turn a metric label into a CSV-safe column suffix (e.g. 'Gross TVPI' -> 'gross_tvpi')."""
    if not label or not str(label).strip():
        return "other"
    s = re.sub(r"[^\w\s]", "", str(label).lower())
    s = re.sub(r"\s+", "_", s).strip("_")
    return s or "other"


def write_csv(rows: List[Dict[str, Any]], output_path: str) -> None:
    """Write CSV: static columns first, then one column set per metric (metric_gt, metric_pred, metric_match). New metrics (e.g. TVPI) append columns."""
    if not rows:
        return
    # Fixed metric columns (always present)
    base_metrics = ["irr", "moic", "dpi"]
    # Collect any other metric names from rows (e.g. tvpi, revenue_growth) for extra columns
    other_metric_names = set()
    for row in rows:
        lbl = row.get("other_metric_label_gt") or ""
        if lbl and str(lbl).strip():
            other_metric_names.add(_sanitize_metric_column_name(lbl))
    other_metric_names = sorted(other_metric_names)
    # Column order: static, then each metric's _gt, _pred, _match (fixed first, then dynamic)
    static = [
        "test_case_id", "model_id", "model_display_name", "provider",
        "input_tokens", "output_tokens", "cost_usd", "parse_ok", "parse_error", "overall_score", "overall_score_is_na",
        "no_metrics_in_scope",
        "irr_in_scope", "moic_in_scope", "dpi_in_scope", "other_in_scope",
        "grading_explanation", "raw_response_truncated", "numbers_extracted_preview",
    ]
    metric_cols = []
    for m in base_metrics + list(other_metric_names):
        metric_cols.extend([f"{m}_gt", f"{m}_pred", f"{m}_match"])
    fieldnames = static + metric_cols

    def flatten_row(row: Dict[str, Any]) -> Dict[str, Any]:
        out = {k: row.get(k, "") for k in static}
        out["raw_response_truncated"] = (row.get("raw_response") or "")[:2000]
        for m in base_metrics:
            out[f"{m}_gt"] = row.get(f"{m}_gt", "")
            out[f"{m}_pred"] = row.get(f"{m}_pred", "")
            out[f"{m}_match"] = row.get(f"{m}_match", "")
        for m in other_metric_names:
            out[f"{m}_gt"] = ""
            out[f"{m}_pred"] = ""
            out[f"{m}_match"] = ""
        # Fill in "other" metric into the right column when this row has one
        other_label = row.get("other_metric_label_gt") or ""
        if other_label:
            col = _sanitize_metric_column_name(other_label)
            if col in other_metric_names:
                out[f"{col}_gt"] = row.get("other_metric_value_gt", "")
                out[f"{col}_pred"] = row.get("other_metric_value_pred", "")
                out[f"{col}_match"] = row.get("other_metric_match", "")
        return out

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(flatten_row(r) for r in rows)
    print(f"Wrote {len(rows)} rows to {output_path}")


def write_trace_json(rows: List[Dict[str, Any]], trace_path: str) -> None:
    """
    Write full LLM trace: one JSON array entry per (test_case, model) with exactly what the LLM
    returned (raw_response, parsed_data) and why the eval graded it (grading_explanation,
    numbers_extracted, per-metric GT/pred/match). Use this for debugging and auditing, not averages.
    """
    if not rows:
        return
    trace_entries = []
    for r in rows:
        entry = {
            "test_case_id": r.get("test_case_id"),
            "model_id": r.get("model_id"),
            "model_display_name": r.get("model_display_name"),
            "provider": r.get("provider"),
            "input_tokens": r.get("input_tokens"),
            "output_tokens": r.get("output_tokens"),
            "cost_usd": r.get("cost_usd"),
            "parse_ok": r.get("parse_ok"),
            "parse_error": r.get("parse_error") or "",
            "overall_score": r.get("overall_score"),
            "overall_score_is_na": r.get("overall_score_is_na"),
            "score_denom": (1 if r.get("irr_in_scope") else 0) + (1 if r.get("moic_in_scope") else 0) + (1 if r.get("dpi_in_scope") else 0) + (1 if r.get("other_in_scope") else 0),
            "no_metrics_in_scope": r.get("no_metrics_in_scope"),
            "grading_explanation": r.get("grading_explanation", ""),
            "raw_response": r.get("raw_response") or "",
            "parsed_data": r.get("parsed_data") or {},
            "numbers_extracted": r.get("numbers_extracted") or [],
            "irr_gt": r.get("irr_gt"), "irr_pred": r.get("irr_pred"), "irr_match": r.get("irr_match"),
            "moic_gt": r.get("moic_gt"), "moic_pred": r.get("moic_pred"), "moic_match": r.get("moic_match"),
            "dpi_gt": r.get("dpi_gt"), "dpi_pred": r.get("dpi_pred"), "dpi_match": r.get("dpi_match"),
            "other_metric_label_gt": r.get("other_metric_label_gt"), "other_metric_value_gt": r.get("other_metric_value_gt"),
            "other_metric_value_pred": r.get("other_metric_value_pred"), "other_metric_match": r.get("other_metric_match"),
        }
        trace_entries.append(entry)
    with open(trace_path, "w", encoding="utf-8") as f:
        json.dump(trace_entries, f, indent=2, default=str)
    print(f"Wrote LLM trace ({len(trace_entries)} entries) to {trace_path}")


def run_on_directory(test_cases_dir: str, output_path: str, tolerance: float = 0.01) -> List[Dict[str, Any]]:
    """
    Run eval on every subdir that contains a PDF and ground_truth.json.
    Each subdir name is used as test_case_id.
    """
    root = Path(test_cases_dir)
    if not root.is_dir():
        raise FileNotFoundError(f"Not a directory: {test_cases_dir}")
    all_rows = []
    for subdir in sorted(d for d in root.iterdir() if d.is_dir() and not d.name.startswith(".")):
        gt_file = subdir / "ground_truth.json"
        pdfs = list(subdir.glob("*.pdf"))
        if not gt_file.exists() or not pdfs:
            continue
        pdf_path = str(pdfs[0])
        print(f"Evaluating: {subdir.name}")
        try:
            rows = evaluate_one(pdf_path, ground_truth_path=str(gt_file), test_case_id=subdir.name, tolerance=tolerance)
            all_rows.extend(rows)
        except Exception as e:
            print(f"  Skip {subdir.name}: {e}")
    if all_rows:
        write_csv(all_rows, output_path)
        trace_path = str(Path(output_path).with_suffix("")) + "_trace.json"
        write_trace_json(all_rows, trace_path)
    return all_rows


def _normalize_for_match(s: str) -> str:
    """Lowercase, keep alphanumeric and spaces, collapse multiple spaces."""
    t = "".join(c if c.isalnum() or c.isspace() else " " for c in s.lower()).strip()
    return " ".join(t.split())


def _match_pdf_to_ground_truth(
    pdf_name: str, pdf_stem: str,
    gt_lookup: Dict[str, Dict[str, Optional[float]]],
    gtpdf_values: List[Tuple[str, str, Dict[str, Optional[float]]]],
) -> Optional[Dict[str, Optional[float]]]:
    """Return ground truth dict for this PDF, or None. Exact (normalized) first, then substring, then word match."""
    gt = gt_lookup.get(pdf_name) or gt_lookup.get(pdf_stem)
    if gt is not None:
        return gt
    pdf_name_norm = _normalize_for_match(pdf_name)
    pdf_stem_norm = _normalize_for_match(pdf_stem)
    pdf_lower = pdf_stem.lower()
    pdf_norm = _normalize_for_match(pdf_stem)
    for excel_key, excel_stem, metrics in gtpdf_values:
        key_norm = _normalize_for_match(excel_key)
        stem_norm = _normalize_for_match(excel_stem)
        # Exact match after normalizing (copy-pasted names, with/without .pdf, extra spaces)
        if key_norm == pdf_name_norm or key_norm == pdf_stem_norm:
            return metrics
        if stem_norm == pdf_name_norm or stem_norm == pdf_stem_norm:
            return metrics
        if len(excel_stem) < 3:
            continue
        key_lower = excel_key.lower()
        stem_lower = excel_stem.lower()
        # Substring match
        if stem_lower in pdf_lower or pdf_lower in stem_lower:
            return metrics
        if key_lower in pdf_lower or pdf_lower in key_lower:
            return metrics
        # Word match: every significant word from Excel (len >= 3) appears in PDF stem
        excel_words = [w for w in key_norm.split() if len(w) >= 3]
        if excel_words and all(w in pdf_norm for w in excel_words):
            return metrics
    return None


def run_on_pdf_dir_with_ground_truth(
    pdf_dir: str,
    output_path: str,
    gt_lookup: Dict[str, Dict[str, Optional[float]]],
    gtpdf_values: List[Tuple[str, str, Dict[str, Optional[float]]]],
    ordered_metrics: List[Dict[str, Optional[float]]],
    tolerance: float = 0.01,
    sample: Optional[int] = None,
    match_by_order: bool = False,
    fail_fast_parse_rate: Optional[float] = 0.98,
    test_case_ids: Optional[List[str]] = None,
    metrics_in_scope: Optional[List[str]] = None,
    debug_dump_text: bool = False,
    allow_projected_metrics: bool = False,
) -> List[Dict[str, Any]]:
    """Run eval on PDFs in pdf_dir using pre-loaded ground truth (from Excel or JSON)."""
    pdf_path = Path(pdf_dir)
    if not pdf_path.is_dir():
        raise FileNotFoundError(f"Not a directory: {pdf_dir}")
    pdf_files = sorted(pdf_path.glob("*.pdf"))
    n_total = len(pdf_files)

    if match_by_order:
        n_pairs = min(len(pdf_files), len(ordered_metrics))
        candidates = [(pdf_files[i], ordered_metrics[i]) for i in range(n_pairs)]
        print(f"Evaluating {n_pairs} PDF(s): match by row position (1st PDF ↔ 1st Excel row).")
        if len(ordered_metrics) < n_total:
            print(f"  Excel has {len(ordered_metrics)} rows but folder has {n_total} PDFs.")
    else:
        candidates = []
        unmatched = []
        for pdf_file in pdf_files:
            gt = _match_pdf_to_ground_truth(pdf_file.name, pdf_file.stem, gt_lookup, gtpdf_values)
            if gt is not None:
                candidates.append((pdf_file, gt))
            else:
                unmatched.append(pdf_file.name)
        print(f"Matching by GTPDF column: found {len(candidates)} PDF(s) with ground truth (out of {n_total} in folder).")
        if unmatched and gtpdf_values:
            print(f"  Add GTPDF rows for the remaining {len(unmatched)} PDFs to evaluate them.")
            if len(unmatched) <= 10:
                print(f"  Unmatched PDFs: {unmatched}")
            else:
                print(f"  Unmatched PDFs (first 10): {unmatched[:10]} ...")

    if test_case_ids:
        allowed = {s.strip() for s in test_case_ids if s and str(s).strip()}
        if allowed:
            candidates = [(p, gt) for p, gt in candidates if p.stem in allowed or p.name in allowed]
            print(f"Filtered to {len(candidates)} PDF(s) matching --test-case-ids.")
    if sample is not None and sample > 0:
        n_with_gt = len(candidates)
        k = min(sample, n_with_gt)
        candidates = random.sample(candidates, k)
        print(f"With --sample {sample}, running on {k} random sample(s): {[p.name for p, _ in candidates]}")
        if k < sample:
            print(f"  Only {n_with_gt} PDF(s) have ground truth; --sample {sample} requested but not enough matches.")
    elif not match_by_order:
        candidates = sorted(candidates, key=lambda x: x[0].name)
    if metrics_in_scope:
        print(f"Scoring restricted to metrics_in_scope: {metrics_in_scope}")
    all_rows = []
    for pdf_file, gt in candidates:
        name = pdf_file.name
        stem = pdf_file.stem
        print(f"Evaluating: {name}")
        try:
            rows = evaluate_one(
                str(pdf_file),
                ground_truth_dict=gt,
                test_case_id=stem,
                tolerance=tolerance,
                metrics_in_scope=metrics_in_scope,
                debug_dump_text=debug_dump_text,
                allow_projected_metrics=allow_projected_metrics,
            )
            all_rows.extend(rows)
        except Exception as e:
            print(f"  Skip {name}: {e}")
    if all_rows:
        write_csv(all_rows, output_path)
        trace_path = str(Path(output_path).with_suffix("")) + "_trace.json"
        write_trace_json(all_rows, trace_path)
        parse_ok_count = sum(1 for r in all_rows if r.get("parse_ok"))
        parse_ok_rate = parse_ok_count / len(all_rows)
        print(f"Run summary: total rows={len(all_rows)}, parse_ok={parse_ok_count}, parse_ok_rate={parse_ok_rate:.2%}")
        if fail_fast_parse_rate is not None and parse_ok_rate < fail_fast_parse_rate:
            raise SystemExit(
                f"parse_ok rate {parse_ok_rate:.2%} is below threshold {fail_fast_parse_rate:.2%}; failing fast."
            )
    return all_rows


def run_on_pdf_dir_with_excel(
    pdf_dir: str,
    excel_path: str,
    output_path: str,
    tolerance: float = 0.01,
    sample: Optional[int] = None,
    match_by_order: bool = False,
    fail_fast_parse_rate: Optional[float] = 0.98,
    test_case_ids: Optional[List[str]] = None,
    metrics_in_scope: Optional[List[str]] = None,
    debug_dump_text: bool = False,
    allow_projected_metrics: bool = False,
) -> List[Dict[str, Any]]:
    """Run eval on PDFs using ground truth from Excel (one tab per doc or GTPDF column)."""
    gt_lookup, gtpdf_values, ordered_metrics = load_ground_truth_from_excel(excel_path)
    return run_on_pdf_dir_with_ground_truth(
        pdf_dir, output_path, gt_lookup, gtpdf_values, ordered_metrics,
        tolerance=tolerance, sample=sample, match_by_order=match_by_order,
        fail_fast_parse_rate=fail_fast_parse_rate,
        test_case_ids=test_case_ids,
        metrics_in_scope=metrics_in_scope,
        debug_dump_text=debug_dump_text,
        allow_projected_metrics=allow_projected_metrics,
    )


def run_on_pdf_dir_with_json(
    pdf_dir: str,
    json_path: str,
    output_path: str,
    tolerance: float = 0.01,
    sample: Optional[int] = None,
    match_by_order: bool = False,
    fail_fast_parse_rate: Optional[float] = 0.98,
    test_case_ids: Optional[List[str]] = None,
    metrics_in_scope: Optional[List[str]] = None,
    debug_dump_text: bool = False,
    allow_projected_metrics: bool = False,
) -> List[Dict[str, Any]]:
    """Run eval on PDFs using ground truth from a multi-document JSON (GTPDF + metrics per entry)."""
    gt_lookup, gtpdf_values, ordered_metrics = load_ground_truth_from_json(json_path)
    return run_on_pdf_dir_with_ground_truth(
        pdf_dir, output_path, gt_lookup, gtpdf_values, ordered_metrics,
        tolerance=tolerance, sample=sample, match_by_order=match_by_order,
        fail_fast_parse_rate=fail_fast_parse_rate,
        test_case_ids=test_case_ids,
        metrics_in_scope=metrics_in_scope,
        debug_dump_text=debug_dump_text,
        allow_projected_metrics=allow_projected_metrics,
    )


def main():
    import argparse
    logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")
    parser = argparse.ArgumentParser(
        description="Metrics-only eval: extract Net IRR, MOIC, DPI; compare to ground truth; output CSV with token/cost."
    )
    parser.add_argument("--pdf", help="Path to a single PDF (or use --dir / --pdf-dir for batch)")
    parser.add_argument("--dir", dest="test_cases_dir", help="Path to directory of test cases (each subdir: PDF + ground_truth.json)")
    parser.add_argument("--pdf-dir", help="Path to directory of PDFs; use with --ground-truth-excel or --ground-truth-json")
    parser.add_argument("--ground-truth", help="Path to ground truth JSON (for single --pdf)")
    parser.add_argument("--ground-truth-excel", dest="ground_truth_excel", help="Path to Excel with expected answers (use with --pdf-dir)")
    parser.add_argument("--ground-truth-json", dest="ground_truth_json", help="Path to multi-document ground truth JSON: { \"Doc\": { \"metrics\": { \"GTPDF\": \"file.pdf\", ... } } } (use with --pdf-dir)")
    parser.add_argument("--output", "-o", default="eval_metrics_results.csv", help="Output CSV path")
    parser.add_argument("--test-case-id", default=None, help="Test case ID for CSV (default: PDF stem)")
    parser.add_argument("--test-case-ids", nargs="*", dest="test_case_ids", default=None, metavar="ID", help="With --pdf-dir: only run on PDFs whose stem or filename matches one of these IDs (e.g. --test-case-ids 'Palmer Square...' 'Coatue Asia...')")
    parser.add_argument("--tolerance", type=float, default=0.01, help="Numeric match tolerance (default 0.01)")
    parser.add_argument("--sample", type=int, default=None, metavar="N", help="Run on N random PDFs only (with --pdf-dir)")
    parser.add_argument("--match-by-order", action="store_true", dest="match_by_order", help="Pair PDFs with Excel by row position (1st PDF ↔ 1st row). Default is to match by GTPDF column (filename).")
    parser.add_argument("--no-fail-fast-parse", action="store_true", dest="no_fail_fast_parse", help="Do not exit when parse_ok rate is below 0.98.")
    parser.add_argument("--metrics-in-scope", dest="metrics_in_scope", default=None, metavar="LIST", help="Comma-separated list of metrics to score (e.g. irr,moic,dpi,current_yield). Default: all GT metrics.")
    parser.add_argument("--debug-dump-text", action="store_true", dest="debug_dump_text", help="Write extracted PDF text to debug_text/<test_case_id>.txt and print lines containing irr|moic|dpi|tvpi|yield|bps|ytd.")
    parser.add_argument("--allow-projected-metrics", action="store_true", dest="allow_projected_metrics", help="If set, score metrics from documents labeled estimate/projection/forecast. Default: exclude them from score_denom.")
    args = parser.parse_args()
    fail_fast_parse_rate: Optional[float] = None if getattr(args, "no_fail_fast_parse", False) else 0.98
    metrics_in_scope: Optional[List[str]] = None
    if getattr(args, "metrics_in_scope", None):
        metrics_in_scope = [s.strip() for s in args.metrics_in_scope.split(",") if s.strip()]
    debug_dump_text: bool = getattr(args, "debug_dump_text", False)
    allow_projected_metrics: bool = getattr(args, "allow_projected_metrics", False)
    if args.test_cases_dir:
        if args.pdf or args.ground_truth or args.pdf_dir or args.ground_truth_excel or getattr(args, "ground_truth_json", None):
            parser.error("Do not mix --dir with --pdf/--ground-truth or --pdf-dir/--ground-truth-excel/--ground-truth-json")
        run_on_directory(args.test_cases_dir, args.output, tolerance=args.tolerance)
        return
    if args.pdf_dir and args.ground_truth_excel:
        if args.pdf or args.ground_truth or getattr(args, "ground_truth_json", None):
            parser.error("Do not mix --pdf-dir with both --ground-truth-excel and --ground-truth-json")
        run_on_pdf_dir_with_excel(
            args.pdf_dir, args.ground_truth_excel, args.output,
            tolerance=args.tolerance, sample=args.sample,
            match_by_order=getattr(args, "match_by_order", False),
            fail_fast_parse_rate=fail_fast_parse_rate,
            test_case_ids=getattr(args, "test_case_ids", None),
            metrics_in_scope=metrics_in_scope,
            debug_dump_text=debug_dump_text,
            allow_projected_metrics=allow_projected_metrics,
        )
        return
    if args.pdf_dir and getattr(args, "ground_truth_json", None):
        if args.pdf or args.ground_truth:
            parser.error("Do not mix --pdf-dir/--ground-truth-json with --pdf/--ground-truth")
        run_on_pdf_dir_with_json(
            args.pdf_dir, args.ground_truth_json, args.output,
            tolerance=args.tolerance, sample=args.sample,
            match_by_order=getattr(args, "match_by_order", False),
            fail_fast_parse_rate=fail_fast_parse_rate,
            test_case_ids=getattr(args, "test_case_ids", None),
            metrics_in_scope=metrics_in_scope,
            debug_dump_text=debug_dump_text,
            allow_projected_metrics=allow_projected_metrics,
        )
        return
    if args.pdf and args.ground_truth:
        test_case_id = args.test_case_id or Path(args.pdf).stem
        print(f"Running metrics eval: PDF={args.pdf}, ground_truth={args.ground_truth}")
        rows = evaluate_one(
            args.pdf,
            ground_truth_path=args.ground_truth,
            test_case_id=test_case_id,
            tolerance=args.tolerance,
            metrics_in_scope=metrics_in_scope,
            debug_dump_text=debug_dump_text,
            allow_projected_metrics=allow_projected_metrics,
        )
        write_csv(rows, args.output)
        trace_path = str(Path(args.output).with_suffix("")) + "_trace.json"
        write_trace_json(rows, trace_path)
        return rows
    parser.error(
        "Provide one of: (1) --pdf and --ground-truth, (2) --dir, or (3) --pdf-dir with --ground-truth-excel or --ground-truth-json"
    )


if __name__ == "__main__":
    main()
